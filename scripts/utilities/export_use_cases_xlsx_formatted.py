from __future__ import annotations

from pathlib import Path
import csv
from textwrap import dedent
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins


def read_use_cases(csv_path: Path) -> list[dict]:
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


def to_numbered_lines(value: str) -> str:
    if not value:
        return ""
    # Split by semicolon separators we used when exporting
    parts = [p.strip() for p in value.split(";") if p.strip()]
    return "\n".join(f"{i+1}. {p}" for i, p in enumerate(parts))


def style_table(ws, table_start_row: int = 2):
    # Column widths
    ws.column_dimensions["A"].width = 24
    # 57.67 ≈ 526 像素
    ws.column_dimensions["B"].width = 57.67

    # Fonts
    base_font = Font(name="Times New Roman", size=12)
    bold = Font(name="Times New Roman", size=12, bold=True)

    # Alignments
    left_col_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    right_col_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # Borders
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    for r in range(table_start_row, max_row + 1):
        a = ws.cell(row=r, column=1)
        b = ws.cell(row=r, column=2)
        a.font = bold
        b.font = base_font
        a.alignment = left_col_alignment
        b.alignment = right_col_alignment
        a.border = border
        b.border = border


def write_use_case_sheet(wb: Workbook, uc: dict):
    uc_id = uc.get("ID", "UC")
    name = uc.get("Name", "")
    title = f"{uc_id}: {name}".strip()
    sheet_name = uc_id
    # Ensure unique sheet name (Excel max 31 chars, unique constraint)
    if sheet_name in wb.sheetnames:
        idx = 2
        while f"{sheet_name}-{idx}" in wb.sheetnames:
            idx += 1
        sheet_name = f"{sheet_name}-{idx}"

    ws = wb.create_sheet(sheet_name)

    # Title row (merged across A:B)
    title_row = 1
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
    title_cell = ws.cell(row=title_row, column=1, value=title)
    title_cell.font = Font(name="Times New Roman", size=12, bold=True)
    # 与示例一致：标题左对齐
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    # Table starts immediately at next row (no spacer)
    start_row = 2

    rows = [
        ("Use Case", title),
        ("Description", uc.get("Description", "")),
        ("Actors", uc.get("Actors", "")),
        ("Preconditions", uc.get("Preconditions", "")),
        ("Postconditions", uc.get("Postconditions", "")),
        ("Standard Process", to_numbered_lines(uc.get("StandardProcess", ""))),
        ("Alternative Process", to_numbered_lines(uc.get("AlternativeProcess", ""))),
    ]

    row_idx = start_row
    for left, right in rows:
        ws.cell(row=row_idx, column=1, value=left)
        ws.cell(row=row_idx, column=2, value=right)
        row_idx += 1

    # 应用表格样式，从第2行开始，确保标题无边框
    style_table(ws, table_start_row=start_row)
    # 页面设置：A4、页边距左右1英寸(2.54cm)、按页宽适配
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=1.0, right=1.0, top=1.0, bottom=1.0)


def main():
    csv_path = Path("../docs/use_cases/UseCases.csv")
    if not csv_path.exists():
        raise SystemExit("UseCases.csv not found. Please run scripts/export_use_cases_csv.py first.")

    use_cases = read_use_cases(csv_path)
    wb = Workbook()
    # Remove default sheet created by openpyxl
    wb.remove(wb.active)

    for uc in use_cases:
        write_use_case_sheet(wb, uc)

    base = Path("../docs/use_cases/UseCases_formatted.xlsx")
    out_path = base
    if out_path.exists():
        idx = 2
        while True:
            candidate = Path(f"../docs/use_cases/UseCases_formatted_v{idx}.xlsx")
            if not candidate.exists():
                out_path = candidate
                break
            idx += 1
    wb.save(out_path)
    print(f"Wrote {out_path.resolve()} with {len(use_cases)} sheets")


if __name__ == "__main__":
    main()


